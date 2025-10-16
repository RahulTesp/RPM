import os
import time
import random
import pytest
import re
import json
from faker import Faker
from datetime import datetime
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# ============================
# Paths & Constants
# ============================
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
RESOURCES_DIR = os.path.join(CURRENT_DIR, "resources")
IMAGE_FILE_PATH = os.path.join(RESOURCES_DIR, "profile.jpg")
EXCEL_FILE_PATH = os.path.join(RESOURCES_DIR, "patient_details_updated.xlsx")
BASE_URL = "https://cx-dev-client.azurewebsites.net/login"
fake = Faker()


# ============================
# Excel Loader
# ============================
def load_excel_data(filepath):
    """Load patient mappings from Excel"""
    wb = load_workbook(filepath)
    sheet = wb.active
    data = []

    # Get header row
    headers = [cell.value for cell in sheet[1]]

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = dict(zip(headers, row))

        # Validation
        if not row_dict.get("patient_email") or not row_dict.get("patient_mobile"):
            raise ValueError(
                f"❌ Missing Email/Mobile in Excel at row {row_num}. "
                f"Email={row_dict.get('patient_email')}, Mobile={row_dict.get('patient_mobile')}"
            )

        data.append({
            "ProgramName": str(row_dict.get("program", "")).strip(),
            "VitalCondition": str(row_dict.get("vital", "")).strip(),
            "Diagnosis": str(row_dict.get("diagnosis", "")).strip(),
            "StateID": str(row_dict.get("state_id", "")).strip(),
            "CityName": str(row_dict.get("city_name", "")).strip(),
            "TimezoneID": str(row_dict.get("timezone_id", "")).strip(),
            "PatientEmail": str(row_dict.get("patient_email", "")).strip(),
            "PatientMobile": str(row_dict.get("patient_mobile", "")).strip(),
            "ClinicCode": str(row_dict.get("clinic_code", "")).strip(),
            "Zipcode": str(row_dict.get("zip_code", "")).strip(),
            "Emergency1": str(row_dict.get("emergency1", "")).strip(),
            "Emergency2": str(row_dict.get("emergency2", "")).strip(),
            "Date": str(row_dict.get("date", "")).strip(),
        })
    return data

PROGRAM_MAPPINGS = load_excel_data(EXCEL_FILE_PATH)

# ============================
# Data Generators
# ============================
def generate_random_patient(program):
    """Generate random patient details (name, DOB, etc.)"""
    feet = random.randint(4, 6)
    inches = random.randint(0, 11)
    weight = random.randint(40, 120)

    return {
        "Firstname": fake.first_name(),
        "Middlename": fake.first_name(),
        "Lastname": fake.last_name(),
        "Email": program["PatientEmail"],
        "MobileNumber": program["PatientMobile"],  # from Excel
        "Date": program["Date"], # from Excel
        "DOB": fake.date_of_birth(minimum_age=18, maximum_age=80).strftime("%Y-%m-%d"),
        "HeightFeet": str(feet),
        "HeightInches": str(inches),
        "Weight": str(weight),
        "WeightUnit": random.choice(["Pound", "Kg"]),
        "Gender": random.choice(["M", "F"]),
        "AddressLine1": fake.street_address(),
        "Notes": fake.sentence(),
    }

def generate_emergency_contacts():
    """Random names & relations"""
    relations = ["Father", "Mother"]
    return {
        "Emergency1Name": fake.first_name(),
        "Emergency1Relation": random.choice(relations),
        "Emergency2Name": fake.first_name(),
        "Emergency2Relation": random.choice(relations),
    }

# ============================
# Credentials Loader
# ============================
def load_credentials():
    username = os.getenv("APP_USERNAME")
    password = os.getenv("APP_PASSWORD")

    if username and password:
        return username, password

    try:
        with open("credentials.txt", "r") as file:
            lines = file.read().splitlines()
            return lines[0].strip(), lines[1].strip()
    except FileNotFoundError:
        raise RuntimeError("No credentials found in environment or credentials.txt file")

# ============================
# Pytest Fixtures
# ============================
@pytest.fixture(scope="session")
def browser_context():
    username, password = load_credentials()

    with sync_playwright() as playwright:
        is_ci = os.getenv("CI") == "true"

        browser = playwright.chromium.launch(
            headless=is_ci,
            slow_mo=500 if not is_ci else 0,
            args=["--start-maximized"] if not is_ci else []
        )

        context = browser.new_context(no_viewport=not is_ci)
        page = context.new_page()

        # Login
        page.goto(BASE_URL, wait_until="networkidle")
        page.fill("#login_username", username)
        page.fill("#login_password", password)
        page.fill("#login_password", password)
        page.wait_for_selector("#btn_login:enabled", timeout=15000)
        page.click("#btn_login")
        time.sleep(30)
        page.wait_for_selector("#homeaddpatient", timeout=30000)

        yield page
        browser.close()


# ============================
# Test Cases
# ============================
@pytest.mark.sanity
@pytest.mark.regression
@pytest.mark.parametrize("program", PROGRAM_MAPPINGS)
def test_add_patient(browser_context, program):
    """
    Adding New patient details
    :param browser_context: Launch playwright server
    :param program: Vital programs
    :return:
    """
    page = browser_context
    form_data = generate_random_patient(program)
    # Navigate
    page.click("#homeaddpatient")
    time.sleep(30)
    page.wait_for_selector("#addPatientFirstname", timeout=10000)
    # Fill patient form
    page.fill("#addPatientFirstname", form_data["Firstname"])
    page.fill("#addPatientMiddlename", form_data["Middlename"])
    page.fill('//*[@id="addPatientlastname"]', form_data["Lastname"])
    page.fill('//*[@id="addPatientemail"]', form_data["Email"])
    page.fill("#addPatientMobile", form_data["MobileNumber"])
    page.fill("#addPatientDOB", form_data["DOB"])
    # Height
    page.eval_on_selector(
        "#addPatientheight",
        """(el, value) => {
            el.value = value;
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        }""",
        form_data["HeightFeet"],
    )
    # Weight
    page.fill("input.duration_value", form_data["Weight"])
    # Gender & Clinic
    page.select_option("#addPatientGender", value=form_data["Gender"])
    page.select_option("#addPatientClinicName", value=program["ClinicCode"])
    # Address
    page.fill("#addPatientAddressLine1", form_data["AddressLine1"])
    page.select_option("#addPatientState", value=program["StateID"])
    page.select_option("#addPatientCity", value=program["CityName"])
    page.fill("#addPatientZipcode", program["Zipcode"])
    page.select_option("#addPatientTimezone", value=program["TimezoneID"])
    # Emergency contacts
    ec = generate_emergency_contacts()
    # Emergency 1
    page.fill("#addPatientEmergencyContact1", ec["Emergency1Name"])
    time.sleep(5)
    page.wait_for_selector("xpath=//*[@id='addPatientRelation1']", timeout=10000)
    options1 = page.locator("xpath=//*[@id='addPatientRelation1']/option").all_text_contents()
    page.select_option("xpath=//*[@id='addPatientRelation1']", label=ec["Emergency1Relation"])
    page.fill("#addPatientRelation1_mobile", value=program["Emergency1"])
    time.sleep(3)
    # Emergency 2
    page.fill("#addPatientEmergencyContact2", ec["Emergency2Name"])
    time.sleep(5)
    page.wait_for_selector("xpath=//*[@id='addPatientRelation2']", timeout=10000)
    options2 = page.locator("xpath=//*[@id='addPatientRelation2']/option").all_text_contents()
    page.select_option("xpath=//*[@id='addPatientRelation2']", label=ec["Emergency2Relation"])
    time.sleep(10)
    page.fill("#addPatientRelation2_mobile", value=program["Emergency2"])
    # Notes
    page.fill("#addPatientAdditionalNotes", form_data["Notes"])
    # Upload image
    if os.path.exists(IMAGE_FILE_PATH):
        page.set_input_files("#image", IMAGE_FILE_PATH)
    # Submit draft
    page.click("#btnaddPatientDraftSubmit")
    time.sleep(10)
    locator = page.locator("h4.subheading-dialog")
    locator.wait_for(state="visible", timeout=5000)
    assert locator.inner_text() == "New Patient added Successfully!"
    # Program & Vital
    page.click("#addPatient")
    time.sleep(3)
    page.select_option("#addPatientProgramName", label=program["ProgramName"])
    time.sleep(5)
    raw_date = program["Date"]
    formatted_date = datetime.strptime(raw_date, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
    page.fill("#addPatientStartDate", formatted_date)
    page.click("#addPatientvitalChange .mat-mdc-select-trigger")
    time.sleep(3)
    page.get_by_text(program["VitalCondition"], exact=True).click()
    time.sleep(3)
    # Diagnosis
    page.get_by_role("button", name="Program Info").click(force=True)
    page.click("input[placeholder='Select Diagnosis']")
    page.get_by_text(program["Diagnosis"], exact=True).click()
    # Confirm
    page.click("#btnaddPatientConfirmSubmit")
    page.locator("button.btn_save", has_text="CONFIRM").click()
    time.sleep(5)
    # Verify
    page.get_by_role("button", name="VIEW PATIENT").click()
    time.sleep(5)
    text = page.locator("span.status_display.patient_prescribed").first.text_content()
    assert text.strip() == "Prescribed"
    patient_name_text = page.locator(
        "xpath=/html/body/app-root/app-dashboard/mat-drawer-container/mat-drawer-content/app-patient-detail-page/div/div/div[1]/div[2]/div[2]/div[2]/div/div/div[1]/div/div[1]/div[3]/p[1]"
    ).text_content().strip()
    expected_name = f"{form_data['Firstname']} {form_data['Lastname']}"
    assert expected_name == patient_name_text, f"Expected '{expected_name}', but got '{patient_name_text}'"
    menu_toggle = page.locator('xpath=//*[@id="screenSmallToggle"]/img')
    if menu_toggle.is_visible():
        menu_toggle.click()
        page.wait_for_timeout(1000)

    # # Click Home menu item ONCE
    page.locator(
        "xpath=/html/body/app-root/app-dashboard/mat-drawer-container/mat-drawer/div/app-side-bar/div/ul/app-profile-menu-button/div[2]/div/div[1]/span").click()
    time.sleep(10)
    page.click('//*[@id="myProfBackHome"]/img')
    time.sleep(10)
